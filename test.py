# -*- coding: utf-8 -*-
import re
import sqlite3
import docx
from io import BytesIO
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Alignment

WORD_FILE = "题目.docx"
EXCEL_FILE = "题目完整.xlsx"
DB_FILE = "题库.db"

def extract_paragraphs_with_images(doc):
    paras = []
    for para in doc.paragraphs:
        text = para.text.strip()
        images = []
        for run in para.runs:
            blips = run.element.findall('.//{http://schemas.openxmlformats.org/drawingml/2006/main}blip')
            for blip in blips:
                rId = blip.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed')
                if rId:
                    img_part = para.part.related_parts[rId]
                    images.append(img_part.blob)
        if text or images:
            paras.append((text, images))
    return paras

def clean_answer_mark(text):
    """删除文本末尾括号内的答案字母，并删除紧跟的句号（中英文），保留括号本身"""
    if not text:
        return text
    text = re.sub(r'[（(][A-Z0-9]+[）)]([。.]?)', r'()', text)
    return text.strip()

def parse_questions(paras):
    questions = []
    current_q = None
    q_num = 0

    header_re = re.compile(r'^[\*\s]*【(.*?)】')
    option_re = re.compile(r'^([A-F])[.、]\s*(.*)')
    answer_re = re.compile(r'^答案[：:]\s*(.*)')
    difficulty_re = re.compile(r'^难易程度[：:]\s*(.*)')
    explanation_re = re.compile(r'^答案解析[：:]\s*(.*)')
    knowledge_re = re.compile(r'^知识点[：:]\s*(.*)')

    for text, images in paras:
        if not text and not images:
            continue

        m = header_re.match(text)
        if m:
            if current_q:
                current_q['题干'] = clean_answer_mark(current_q['题干'])
                questions.append(current_q)
            q_num += 1
            q_type = m.group(1)
            stem = text[len(m.group(0)):].strip()
            stem = clean_answer_mark(stem)
            current_q = {
                '题号': str(q_num), '题型': q_type, '题干': stem,
                '题干图片': None, '解析图片': None,
                'A': '', 'B': '', 'C': '', 'D': '', 'E': '', 'F': '',
                '答案': '', '难易': '中', '解析': '', '知识点': ''
            }
            if images:
                current_q['题干图片'] = images[0]
            continue

        if current_q is None:
            continue

        lines = text.split('\n')
        for line in lines:
            line = line.strip()
            if not line:
                continue

            opt = option_re.match(line)
            if opt:
                letter = opt.group(1)
                content = opt.group(2).strip()
                content = clean_answer_mark(content)
                current_q[letter] = content
                continue

            ans = answer_re.match(line)
            if ans:
                current_q['答案'] = ans.group(1).strip()
                continue

            diff = difficulty_re.match(line)
            if diff:
                current_q['难易'] = diff.group(1).strip()
                continue

            exp = explanation_re.match(line)
            if exp:
                current_q['解析'] = exp.group(1).strip()
                continue

            know = knowledge_re.match(line)
            if know:
                current_q['知识点'] = know.group(1).strip()
                continue

            if current_q['题干']:
                current_q['题干'] += ' ' + line
            else:
                current_q['题干'] = line

        if images and current_q and current_q['解析图片'] is None:
            if current_q['题干图片'] is None or images[0] != current_q['题干图片']:
                current_q['解析图片'] = images[0]

    if current_q:
        current_q['题干'] = clean_answer_mark(current_q['题干'])
        questions.append(current_q)

    return questions

def export_excel(questions):
    wb = Workbook()
    ws = wb.active
    ws.title = "题库"

    col_widths = {'A': 8, 'B': 12, 'C': 60, 'D': 15, 'E': 30, 'F': 30, 'G': 30,
                  'H': 30, 'I': 30, 'J': 30, 'K': 12, 'L': 10, 'M': 50, 'N': 15, 'O': 30}
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    headers = ["题号","题型","题干","题干图片","A","B","C","D","E","F",
               "答案","难易","解析","解析图片","知识点"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for i, q in enumerate(questions, start=2):
        ws.cell(row=i, column=1, value=q["题号"])
        ws.cell(row=i, column=2, value=q["题型"])
        stem_text = q["题干"].replace('\n', ' ').replace('\r', ' ')
        ws.cell(row=i, column=3, value=stem_text)
        ws.cell(row=i, column=5, value=q["A"])
        ws.cell(row=i, column=6, value=q["B"])
        ws.cell(row=i, column=7, value=q["C"])
        ws.cell(row=i, column=8, value=q["D"])
        ws.cell(row=i, column=9, value=q["E"])
        ws.cell(row=i, column=10, value=q["F"])
        ws.cell(row=i, column=11, value=q["答案"])
        ws.cell(row=i, column=12, value=q["难易"])
        ws.cell(row=i, column=13, value=q["解析"])
        ws.cell(row=i, column=15, value=q["知识点"])

        ws.row_dimensions[i].height = 100
        for col in [3, 13]:
            ws.cell(row=i, column=col).alignment = Alignment(wrapText=True)

        if q.get("题干图片"):
            try:
                img = XLImage(BytesIO(q["题干图片"]))
                max_w = 150
                if img.width > max_w:
                    img.height = int(max_w * img.height / img.width)
                    img.width = max_w
                ws.add_image(img, f"D{i}")
            except Exception as e:
                print(f"题号 {q['题号']} 题干图片插入失败: {e}")

        if q.get("解析图片"):
            try:
                img = XLImage(BytesIO(q["解析图片"]))
                max_w = 150
                if img.width > max_w:
                    img.height = int(max_w * img.height / img.width)
                    img.width = max_w
                ws.add_image(img, f"N{i}")
            except Exception as e:
                print(f"题号 {q['题号']} 解析图片插入失败: {e}")

    wb.save(EXCEL_FILE)
    print(f"✅ Excel 导出完成，共 {len(questions)} 题，保存为 {EXCEL_FILE}")

def export_database(questions):
    """将题目数据（含图片二进制）存入 SQLite 数据库"""
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS questions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            question_number TEXT UNIQUE NOT NULL,
            question_type TEXT,
            stem TEXT,
            stem_image BLOB,
            option_a TEXT,
            option_b TEXT,
            option_c TEXT,
            option_d TEXT,
            option_e TEXT,
            option_f TEXT,
            answer TEXT,
            difficulty TEXT,
            explanation TEXT,
            explanation_image BLOB,
            knowledge TEXT
        )
    ''')
    # 清空旧数据（可选）
    cursor.execute('DELETE FROM questions')
    for q in questions:
        cursor.execute('''
            INSERT INTO questions (
                question_number, question_type, stem, stem_image,
                option_a, option_b, option_c, option_d, option_e, option_f,
                answer, difficulty, explanation, explanation_image, knowledge
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            q['题号'], q['题型'], q['题干'], q.get('题干图片'),
            q['A'], q['B'], q['C'], q['D'], q['E'], q['F'],
            q['答案'], q['难易'], q['解析'], q.get('解析图片'), q['知识点']
        ))
    conn.commit()
    conn.close()
    print(f"✅ 数据库导出完成，共 {len(questions)} 题，保存为 {DB_FILE}")

def main():
    try:
        doc = docx.Document(WORD_FILE)
        paras = extract_paragraphs_with_images(doc)
        print(f"共读取 {len(paras)} 个段落，其中包含图片的段落数：{sum(1 for _, im in paras if im)}")
        questions = parse_questions(paras)
        if not questions:
            print("警告：未解析到任何题目，请检查题头格式。前5个段落文本：")
            for i, (text, _) in enumerate(paras[:5]):
                print(f"段落{i+1}: {text[:100]}")
        export_excel(questions)
        export_database(questions)
    except Exception as e:
        print(f"❌ 运行失败：{e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()